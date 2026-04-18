"""链路预算模块 — 表格式界面"""

import math
import io
import numpy as np

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton,
    QFrame, QFileDialog, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QStyledItemDelegate,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QFont, QBrush, QKeyEvent

from ui.base_dialog import ModuleDialog


# ══════════════════════════════════════════════════════════
#  常量
# ══════════════════════════════════════════════════════════

BOLTZMANN_DBW = -228.6

# 调制阶数 -> 标签 / 解调门限
MOD_ORDERS  = [4,      8,      16,      32,      64,      128,     256,     512    ]
MOD_LABELS  = ["QPSK","8PSK","16QAM","32QAM","64QAM","128QAM","256QAM","512QAM"]
DEMOD_THRESH= [6.58,  11.29,  13.13,  16.02,  19.07,  21.42,   24.27,   26.95  ]

def _order_to_label(order_str):
    """将用户输入的调制阶数（如"4"/"QPSK"）转为标准标签"""
    s = str(order_str).strip().upper()
    for o, lbl in zip(MOD_ORDERS, MOD_LABELS):
        if s == str(o) or s == lbl.upper():
            return lbl
    return "QPSK"

def _label_to_thresh(label):
    if label in MOD_LABELS:
        return DEMOD_THRESH[MOD_LABELS.index(label)]
    return DEMOD_THRESH[0]


# ══════════════════════════════════════════════════════════
#  行定义
#  (显示名, 符号, 单位, 类型)
#  类型: "section" 分组标题 | "input" 用户填写 | "calc" 自动计算 | "fixed" 固定常数
# ══════════════════════════════════════════════════════════

ROWS = [
    # 轨道参数
    ("轨道参数",           "",      "",         "section"),
    ("地球半径",           "r",     "km",       "input"),
    ("卫星高度",           "h",     "km",       "input"),
    ("地面站最低仰角",     "θ",     "°",        "input"),
    ("卫星天线最大指向角", "α",     "°",        "calc"),   # 由 h/r/θ 算出
    ("地心夹角",           "γ",     "°",        "calc"),   # 由 h/r/θ 算出
    ("卫星到地心距离",     "h+r",   "km",       "calc"),
    ("信号传输距离",       "d",     "km",       "calc"),
    # 发射端
    ("发射端",             "",      "",         "section"),
    ("等效全向辐射功率",   "EIRP",  "dBW",      "input"),
    # 链路损耗
    ("链路损耗",           "",      "",         "section"),
    ("工作频率",           "f",     "GHz",      "input"),
    ("自由空间损耗",       "Lfs",   "dB",       "calc"),
    ("大气损耗",           "Latm",  "dB",       "input"),
    ("指向误差",           "",      "dB",       "input"),
    ("极化损失",           "",      "dB",       "input"),
    ("雨衰",               "Lr",    "dB",       "input"),
    ("链路总损耗",         "L",     "dB",       "calc"),
    # 接收端
    ("接收端",             "",      "",         "section"),
    ("天线G/T",            "G/T",   "dB/K",     "input"),
    ("玻尔兹曼常数",       "k",     "dBW/Hz/K", "fixed"),
    ("符号速率",           "Rs",    "Mbps",     "input"),
    ("带内噪声",           "BN",    "dBHz",     "calc"),
    ("C/N",                "C/N",   "dB",       "calc"),
    # 链路余量（每种调制模式一行，标题行去掉）
    ("调制模式",           "",      "阶数",     "input"),
    ("编码模式",           "R",     "",         "input"),
    ("解调门限下限",       "",      "dB",       "input"),   # 可手动设置
    ("链路余量",           "Δ",  "dB",       "calc"),
]

# 默认值
DEFAULTS = {
    "地球半径":           "6371",
    "卫星高度":           "508",
    "地面站最低仰角":     "10",
    "等效全向辐射功率":   "44.8",
    "工作频率":           "39",
    "大气损耗":           "3.2",
    "指向误差":           "0.5",
    "极化损失":           "0.5",
    "雨衰":               "0.00",
    "天线G/T":            "25",
    "玻尔兹曼常数":       str(BOLTZMANN_DBW),
    "符号速率":           "750",
    "调制模式":           "4",
    "编码模式":           "7/8",
    "解调门限下限":       "6.58",
}

# 快速索引
RI           = {r[0]: i for i, r in enumerate(ROWS)}
SECTION_ROWS = {i for i, r in enumerate(ROWS) if r[3] == "section"}
INPUT_ROWS   = {i for i, r in enumerate(ROWS) if r[3] == "input"}
CALC_ROWS    = {i for i, r in enumerate(ROWS) if r[3] == "calc"}
FIXED_ROWS   = {i for i, r in enumerate(ROWS) if r[3] == "fixed"}
MARGIN_ROWS  = {RI["链路余量"]}

# 颜色
C_SECTION = QColor("#DCE8F5")
C_INPUT   = QColor("#FFFFFF")
C_CALC    = QColor("#FFFBF0")
C_FIXED   = QColor("#F2F2F0")
C_POS     = QColor("#1A6B35")
C_NEG     = QColor("#B00000")
C_HDR_BG  = "#2E6DB4"


# ══════════════════════════════════════════════════════════
#  计算
# ══════════════════════════════════════════════════════════

def _f(s, d=0.0):
    try:    return float(str(s).strip())
    except: return d

def _fspl(freq_ghz, dist_km):
    if dist_km <= 0 or freq_ghz <= 0: return 0.0
    return 20*math.log10(dist_km) + 20*math.log10(freq_ghz) + 92.45

def _slant_range(h, r, theta_deg):
    if theta_deg <= 0: return 0.0
    t = math.radians(theta_deg)
    return math.sqrt((r+h)**2 - (r*math.cos(t))**2) - r*math.sin(t)

def _central_angle(h, r, theta_deg):
    """地心夹角 γ（卫星-地心-地面站 的夹角，单位°）"""
    t = math.radians(theta_deg)
    arg = max(-1.0, min(1.0, r*math.cos(t)/(r+h)))
    return math.degrees(math.acos(arg)) - theta_deg

def _max_pointing_angle(h, r, theta_deg):
    """
    卫星天线最大指向角 α：
    由正弦定理: sin(α)/r = sin(90°+θ)/(r+h) => sin(α) = r*cos(θ)/(r+h)
    """
    t = math.radians(theta_deg)
    arg = max(-1.0, min(1.0, r*math.cos(t)/(r+h)))
    return math.degrees(math.asin(arg))

def calc_column(vals: dict) -> dict:
    out = dict(vals)

    r     = _f(vals.get("地球半径",          "6371"))
    h     = _f(vals.get("卫星高度",          "508"))
    theta = _f(vals.get("地面站最低仰角",    "10"))
    freq  = _f(vals.get("工作频率",          "39"))
    eirp  = _f(vals.get("等效全向辐射功率",  "44.8"))
    atm   = _f(vals.get("大气损耗",          "3.2"))
    point = _f(vals.get("指向误差",          "0.5"))
    polar = _f(vals.get("极化损失",          "0.5"))
    rain  = _f(vals.get("雨衰",              "0"))
    gt    = _f(vals.get("天线G/T",           "25"))
    rs    = _f(vals.get("符号速率",          "750"))

    mod_raw = vals.get("调制模式", "4").strip()
    mod_lbl = _order_to_label(mod_raw)
    code    = vals.get("编码模式", "7/8").strip()

    # 解调门限：优先用户填写，否则按调制模式查表
    thresh_raw = vals.get("解调门限下限", "").strip()
    thresh = _f(thresh_raw) if thresh_raw else _label_to_thresh(mod_lbl)

    # 几何计算
    d     = _slant_range(h, r, theta)
    gamma = _central_angle(h, r, theta)
    alpha = _max_pointing_angle(h, r, theta)

    # 链路计算
    lfs   = _fspl(freq, d)
    l_tot = lfs + atm + point + polar + rain
    bn    = 10*math.log10(max(rs, 1e-9)*1e6)
    cn    = eirp - l_tot + gt - BOLTZMANN_DBW - bn

    out["卫星天线最大指向角"] = f"{alpha:.2f}"
    out["地心夹角"]           = f"{gamma:.2f}"
    out["卫星到地心距离"]     = f"{r+h:.2f}"
    out["信号传输距离"]       = f"{d:.2f}"
    out["自由空间损耗"]       = f"{lfs:.2f}"
    out["链路总损耗"]         = f"{l_tot:.2f}"
    out["玻尔兹曼常数"]       = str(BOLTZMANN_DBW)
    out["带内噪声"]           = f"{bn:.2f}"
    out["C/N"]                = f"{cn:.2f}"
    # 解调门限如果用户没填，写回查表值
    if not thresh_raw:
        out["解调门限下限"] = f"{thresh:.2f}"

    out["链路余量"] = f"{cn - thresh:.2f}"

    return out


# ══════════════════════════════════════════════════════════
#  自定义委托：编辑时白底深色字，清晰可见
# ══════════════════════════════════════════════════════════

class _EditDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        ed = QLineEdit(parent)
        ed.setStyleSheet(
            "QLineEdit{"
            "  background:#FFFFFF; color:#111111;"
            "  border:2px solid #2E6DB4; border-radius:2px;"
            "  padding:1px 4px; font-size:12px;"
            "  selection-background-color:#BEDAF7;"
            "}")
        return ed

    def setEditorData(self, ed, index):
        ed.setText(index.data(Qt.ItemDataRole.EditRole) or "")
        ed.selectAll()

    def setModelData(self, ed, model, index):
        model.setData(index, ed.text(), Qt.ItemDataRole.EditRole)


# ══════════════════════════════════════════════════════════
#  对话框
# ══════════════════════════════════════════════════════════

class LinkBudgetDialog(ModuleDialog):
    TITLE        = "链路预算"
    SUBTITLE     = ""
    ACCENT_COLOR = "#378ADD"
    MIN_WIDTH    = 900
    MIN_HEIGHT   = 680

    def __init__(self, parent=None):
        self._n_cols   = 5
        self._updating = False
        super().__init__(parent)
        self.resize(1200, 760)

    def keyPressEvent(self, event: QKeyEvent):
        # 屏蔽 Enter / Escape 防止意外关闭
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter,
                           Qt.Key.Key_Escape):
            event.ignore()
            return
        super().keyPressEvent(event)

    def build_content(self, layout: QVBoxLayout):
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        # ── 工具栏 ────────────────────────────────────────
        tb = QHBoxLayout(); tb.setSpacing(6)
        for txt, color, slot in [
            ("＋ 添加列",   "#2E6DB4", self._add_col),
            ("删除选中列",  "#888888", self._del_selected_col),
            ("导出 Excel", "#1D6B42", self._export_excel),
        ]:
            btn = QPushButton(txt)
            btn.setFixedHeight(28)
            btn.setStyleSheet(self._bstyle(color))
            btn.clicked.connect(slot)
            tb.addWidget(btn)
        tb.addStretch()
        layout.addLayout(tb)

        # ── 表格 ──────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background:#FFFFFF; border:1px solid #C8D8EC;
                gridline-color:#D8E4F0; font-size:12px; color:#1A1A1A;
            }}
            QTableWidget::item {{ padding:2px 5px; }}
            QTableWidget::item:selected {{ background:#C5DCF5; color:#1A1A1A; }}
            QHeaderView::section {{
                background:{C_HDR_BG}; color:#FFFFFF;
                font-size:12px; font-weight:600;
                padding:4px 5px; border:none;
                border-right:1px solid #4A88C7;
                border-bottom:1px solid #4A88C7;
            }}
        """)
        self.table.setAlternatingRowColors(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setHorizontalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setVerticalScrollMode(
            QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked |
            QAbstractItemView.EditTrigger.AnyKeyPressed)
        self.table.setItemDelegate(_EditDelegate(self.table))
        self.table.itemChanged.connect(self._on_changed)

        layout.addWidget(self.table, stretch=1)
        self._build_table()

    # ── 表格构建 ──────────────────────────────────────────

    def _build_table(self):
        self._updating = True
        n_rows = len(ROWS)
        n_cols = 3 + self._n_cols

        self.table.setRowCount(n_rows)
        self.table.setColumnCount(n_cols)

        self.table.setHorizontalHeaderLabels(
            ["参数", "符号", "单位"] +
            [f"场景 {i+1}" for i in range(self._n_cols)])

        # 固定列宽（允许拖拽）
        self.table.setColumnWidth(0, 155)
        self.table.setColumnWidth(1, 46)
        self.table.setColumnWidth(2, 60)
        for ci in range(3, n_cols):
            self.table.setColumnWidth(ci, 100)

        # 前三列改为 Interactive（可拖拽），不再 Fixed
        hdr = self.table.horizontalHeader()
        for ci in range(n_cols):
            hdr.setSectionResizeMode(ci, QHeaderView.ResizeMode.Interactive)
        # 最后一列拉伸填满
        hdr.setStretchLastSection(False)

        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            self.table.setRowHeight(ri, 22)
            self._set_meta(ri, 0, name, rtype)
            self._set_meta(ri, 1, sym,  rtype)
            self._set_meta(ri, 2, unit, rtype)
            for ci in range(self._n_cols):
                self._init_data(ri, 3 + ci)

        self._updating = False
        self._calc_all()

    def _set_meta(self, ri, ci, text, rtype):
        item = QTableWidgetItem(text)
        item.setFlags(Qt.ItemFlag.ItemIsEnabled)   # 不可编辑，但可拖拽列宽
        if rtype == "section":
            item.setBackground(QBrush(C_SECTION))
            f = item.font(); f.setBold(True); item.setFont(f)
            item.setForeground(QBrush(QColor("#1A4A80")))
        else:
            item.setBackground(QBrush(QColor("#F5F8FC")))
            item.setForeground(QBrush(QColor("#444444")))
        self.table.setItem(ri, ci, item)

    def _init_data(self, ri, ci):
        name, sym, unit, rtype = ROWS[ri]

        if rtype == "section":
            item = QTableWidgetItem("")
            item.setFlags(Qt.ItemFlag.NoItemFlags)
            item.setBackground(QBrush(C_SECTION))
            self.table.setItem(ri, ci, item)
            return

        default = DEFAULTS.get(name, "")
        item = QTableWidgetItem(default)

        if rtype == "input":
            item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                          Qt.ItemFlag.ItemIsSelectable |
                          Qt.ItemFlag.ItemIsEditable)
            item.setBackground(QBrush(C_INPUT))
        elif rtype == "fixed":
            item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                          Qt.ItemFlag.ItemIsSelectable)
            item.setBackground(QBrush(C_FIXED))
            item.setForeground(QBrush(QColor("#888888")))
        else:  # calc
            item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                          Qt.ItemFlag.ItemIsSelectable)
            item.setBackground(QBrush(C_CALC))

        self.table.setItem(ri, ci, item)

    # ── 增/删列 ───────────────────────────────────────────

    def _add_col(self):
        self._n_cols += 1
        ci = self.table.columnCount()
        self.table.insertColumn(ci)
        self.table.setHorizontalHeaderItem(
            ci, QTableWidgetItem(f"场景 {self._n_cols}"))
        self.table.setColumnWidth(ci, 100)
        self.table.horizontalHeader().setSectionResizeMode(
            ci, QHeaderView.ResizeMode.Interactive)
        self._updating = True
        for ri in range(len(ROWS)):
            self._init_data(ri, ci)
        self._updating = False
        self._calc_col(ci)

    def _del_selected_col(self):
        # 收集选中的数据列（列索引 >= 3），降序排列以便安全删除
        sel_cols = sorted(
            {idx.column() for idx in self.table.selectedIndexes()
             if idx.column() >= 3},
            reverse=True)
        if not sel_cols:
            QMessageBox.information(self, "提示", "请先选中要删除的数据列")
            return
        if self._n_cols - len(sel_cols) < 1:
            QMessageBox.information(self, "提示", "至少保留一列")
            return
        for c in sel_cols:
            self.table.removeColumn(c)
            self._n_cols -= 1
        # 删除后重命名剩余数据列：场景 1, 2, 3 ...
        self._renumber_cols()

    def _renumber_cols(self):
        """删列后将数据列标题重新编号为 场景 1, 2, 3 ..."""
        idx = 1
        for ci in range(3, self.table.columnCount()):
            self.table.setHorizontalHeaderItem(
                ci, QTableWidgetItem(f"场景 {idx}"))
            idx += 1

    # ── 计算 ──────────────────────────────────────────────

    def _on_changed(self, item):
        if self._updating:
            return
        ci = item.column()
        if ci >= 3:
            self._calc_col(ci)

    def _calc_all(self):
        for ci in range(3, self.table.columnCount()):
            self._calc_col(ci)

    def _calc_col(self, ci):
        vals = {}
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if rtype in ("input", "fixed"):
                it = self.table.item(ri, ci)
                vals[name] = it.text() if it else DEFAULTS.get(name, "")
        result = calc_column(vals)

        self._updating = True
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if rtype != "calc":
                continue
            it = self.table.item(ri, ci)
            if it is None:
                it = QTableWidgetItem()
                it.setFlags(Qt.ItemFlag.ItemIsEnabled |
                             Qt.ItemFlag.ItemIsSelectable)
                it.setBackground(QBrush(C_CALC))
                self.table.setItem(ri, ci, it)

            v = result.get(name, "")
            it.setText(v)

            if ri in MARGIN_ROWS:
                try:
                    fv = float(v)
                    it.setForeground(QBrush(C_POS if fv >= 0 else C_NEG))
                    ft = it.font(); ft.setBold(True); it.setFont(ft)
                    it.setBackground(QBrush(
                        QColor("#EBF7EE") if fv >= 0 else QColor("#FDECEA")))
                except:
                    pass
        self._updating = False

    # ── 导出 Excel ────────────────────────────────────────

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel", "链路预算.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            self._write_excel(path)
            QMessageBox.information(self, "导出成功", f"已保存：\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def _write_excel(self, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook(); ws = wb.active; ws.title = "链路预算"
        thin = Side(style="thin", color="C8D8EC")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr  = Alignment(horizontal="center", vertical="center")
        lft  = Alignment(horizontal="left",   vertical="center")

        def fill(h): return PatternFill("solid", start_color=h, end_color=h)

        n_data = self._n_cols
        # 表头
        ws.row_dimensions[1].height = 24
        hdrs = (["参数","符号","单位"] +
                [self.table.horizontalHeaderItem(3+i).text()
                 for i in range(n_data)])
        for ci_x, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci_x, h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = fill("2E6DB4"); c.alignment = ctr; c.border = brd
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 7
        ws.column_dimensions["C"].width = 10
        for i in range(n_data):
            ws.column_dimensions[get_column_letter(4+i)].width = 13

        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            er = ri + 2
            ws.row_dimensions[er].height = 18
            is_sec = (rtype == "section")
            for ci_x, txt in enumerate([name, sym, unit], 1):
                c = ws.cell(er, ci_x, txt)
                c.border = brd
                c.alignment = lft if ci_x == 1 else ctr
                c.font = Font(size=10, bold=is_sec,
                              color="1A4A80" if is_sec else "222222")
                c.fill = fill("DCE8F5") if is_sec else fill("F5F8FC")
            for di in range(n_data):
                it = self.table.item(ri, 3+di)
                raw = it.text() if it else ""
                try:    val = float(raw)
                except: val = raw
                c = ws.cell(er, 4+di, val)
                c.border = brd; c.alignment = ctr
                c.number_format = "0.00"
                if is_sec:
                    c.fill = fill("DCE8F5")
                    c.font = Font(size=10, bold=True, color="1A4A80")
                elif rtype == "fixed":
                    c.fill = fill("F2F2F0")
                    c.font = Font(size=10, color="888888")
                elif rtype == "calc":
                    c.fill = fill("FFFBF0")
                    c.font = Font(size=10)
                    if ri in MARGIN_ROWS and isinstance(val, float):
                        c.font = Font(size=10,
                                      color="1A6B35" if val >= 0 else "B00000",
                                      bold=True)
                        c.fill = fill("EBF7EE" if val >= 0 else "FDECEA")
                else:
                    c.fill = fill("FFFFFF"); c.font = Font(size=10)
        ws.freeze_panes = "A2"
        wb.save(path)

    @staticmethod
    def _bstyle(color):
        from PyQt6.QtGui import QColor as QC
        dark = QC(color).darker(130).name()
        return (f"QPushButton{{background:{color};color:#FFF;border:none;"
                f"border-radius:4px;padding:0 12px;font-size:12px;}}"
                f"QPushButton:hover{{background:{dark};}}")