"""
无线信道建模 — 全部效应集成在单张表格中

两种模式：
  城市模式  — 选择城市后，5 列固定超越概率（0.001%/0.01%/0.1%/1%/5%），
              R_p 由 P.618-14 外推，可输出统计雨衰图像。
  自定义模式 — 列标题为「场景 1」「场景 2」…，可增删列，
              每列降雨强度 R 独立输入，不可输出图像。
"""

import math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QComboBox,
    QDialog, QFileDialog, QMessageBox,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QStyledItemDelegate, QSizePolicy,
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QBrush

from ui.base_dialog import ModuleDialog
from modules.rain_attenuation import (
    compute_rain_statistics,
    compute_rain_attenuation,
    calc_effective_path_length,
)
from modules.atmospheric_attenuation import compute_atm_attenuation
from modules.cloud_attenuation import compute_cloud_attenuation
from modules.ionosphere_effects import calc_ionospheric_group_delay
from modules.cross_polarization import calc_xpd_rain
from modules.itu_data_loader import (
    CITY_NAMES, get_city_rain_params, data_source_info,
)


# ── 字体 ──────────────────────────────────────────────────
def _setup_font():
    for n in ["Microsoft YaHei", "SimHei", "PingFang SC", "Noto Sans CJK SC"]:
        if n in {f.name for f in fm.fontManager.ttflist}:
            plt.rcParams["font.family"] = n
            plt.rcParams["axes.unicode_minus"] = False
            return
_setup_font()


# ══════════════════════════════════════════════════════════
#  城市模式：超越概率列表
# ══════════════════════════════════════════════════════════
PROBS       = [0.001, 0.01, 0.1, 1.0, 5.0]
PROB_LABELS = ["0.001%", "0.01%", "0.1%", "1%", "5%"]
N_PROB      = len(PROBS)


# ══════════════════════════════════════════════════════════
#  行定义（城市/自定义模式共用同一套行）
#
#  rtype：
#    section  — 分组标题
#    input    — 用户可编辑（自定义：各列独立；城市：列间同步）
#    rain_in  — 仅自定义模式下可编辑的降雨强度（城市模式为 calc）
#    coord    — 经纬度，只读，城市模式填入，自定义时为空
#    calc     — 自动计算
# ══════════════════════════════════════════════════════════

ROWS = [
    # ── 链路参数 ──────────────────────────────────────────
    ("链路参数",         "",       "",               "section"),
    ("工作频率",         "f",      "GHz",            "input"),
    ("地面站仰角",       "θ",      "°",              "input"),
    ("极化倾角",         "τ",      "°",              "input"),
    ("地面站海拔",       "h_s",    "km",             "input"),

    # ── 降雨参数 ──────────────────────────────────────────
    ("降雨参数",         "",       "",               "section"),
    ("雨顶高度",         "H_R",    "km",             "input"),
    ("地面站纬度",       "lat",    "°",              "coord"),
    ("地面站经度",       "lon",    "°",              "coord"),

    # ── 雨衰路径 ──────────────────────────────────────────
    ("雨衰路径",         "",       "",               "section"),
    ("有效路径长度",     "L_eff",  "km",             "calc"),

    # ── 降雨强度 ──────────────────────────────────────────
    # 城市模式：calc（由 R_0.01 外推）
    # 自定义模式：rain_in（用户输入）
    ("降雨强度",         "",       "",               "section"),
    ("R",                "R",      "mm/h",           "rain_in"),

    # ── 雨比衰减 ──────────────────────────────────────────
    ("雨比衰减",         "",       "",               "section"),
    ("γ_R",              "γ_R",    "dB/km",          "calc"),

    # ── 雨衰结果 ──────────────────────────────────────────
    ("雨衰结果",         "",       "",               "section"),
    ("链路雨衰",         "A_rain", "dB",             "calc"),

    # ── 大气参数 ──────────────────────────────────────────
    ("大气参数",         "",       "",               "section"),
    ("大气压",           "p",      "hPa",            "input"),
    ("温度",             "T",      "K",              "input"),
    ("水汽分压",         "e",      "hPa",            "input"),

    # ── 大气比衰减 ────────────────────────────────────────
    ("大气比衰减",       "",       "",               "section"),
    ("O₂ 比衰减",        "γ_O₂",   "dB/km",          "calc"),
    ("H₂O 比衰减",       "γ_H₂O",  "dB/km",          "calc"),
    ("总气体比衰减",     "γ_gas",  "dB/km",          "calc"),

    # ── 大气衰减结果 ──────────────────────────────────────
    ("大气衰减结果",     "",       "",               "section"),
    ("O₂ 路径衰减",      "A_O₂",   "dB",             "calc"),
    ("H₂O 路径衰减",     "A_H₂O",  "dB",             "calc"),
    ("总大气气体衰减",   "A_gas",  "dB",             "calc"),

    # ── 云雾参数 ──────────────────────────────────────────
    ("云雾参数",         "",       "",               "section"),
    ("液态水柱含量",     "L",      "kg/m²",          "input"),

    # ── 云雾衰减结果 ──────────────────────────────────────
    ("云雾衰减结果",     "",       "",               "section"),
    ("云雾比衰减系数",   "K_L",    "(dB/km)/(g/m³)", "calc"),
    ("云雾路径衰减",     "A_cloud","dB",             "calc"),

    # ── 综合损耗 ──────────────────────────────────────────
    ("综合损耗",         "",       "",               "section"),
    ("雨衰+大气+云雾",   "A_total","dB",             "calc"),

    # ── 电离层参数 ────────────────────────────────────────
    ("电离层参数",       "",       "",               "section"),
    ("电子总含量 TEC",   "N_T",    "el/m²",          "input"),

    # ── 电离层群时延 ──────────────────────────────────────
    ("电离层群时延",     "",       "",               "section"),
    ("群时延",           "t",      "ns",             "calc"),

    # ── 交叉极化效应 ──────────────────────────────────────
    ("交叉极化效应",     "",       "",               "section"),
    ("雨滴伪角",         "σ",      "°",              "input"),
    ("雨致 XPD",         "XPD_R",  "dB",             "calc"),
]

DEFAULTS = {
    "工作频率":       "39",
    "地面站仰角":     "10",
    "极化倾角":       "45",
    "地面站海拔":     "0.0",
    "雨顶高度":       "4.0",
    "地面站纬度":     "",
    "地面站经度":     "",
    "R":              "30",     # 自定义模式降雨强度默认值
    "大气压":         "1013.25",
    "温度":           "288.15",
    "水汽分压":       "10.0",
    "液态水柱含量":   "0.5",
    "电子总含量 TEC": "1e17",
    "雨滴伪角":       "5",
}

# 行索引
_RI        = {r[0]: i for i, r in enumerate(ROWS)}
_ROW_RAIN  = _RI["链路雨衰"]
_ROW_GAS   = _RI["总大气气体衰减"]
_ROW_CLOUD = _RI["云雾路径衰减"]
_ROW_TOTAL = _RI["雨衰+大气+云雾"]
_ROW_XPD_R = _RI["雨致 XPD"]
_ROW_R     = _RI["R"]           # 降雨强度行

_C_SECTION = QColor("#DCE8F5")
_C_INPUT   = QColor("#FFFFFF")
_C_CALC    = QColor("#FFFBF0")
_C_HDR_BG  = "#2E6B8A"


# ══════════════════════════════════════════════════════════
#  核心计算
# ══════════════════════════════════════════════════════════

def _fv(s, d=0.0):
    try:    return float(str(s).strip())
    except: return d


def _calc_single(vals: dict, rain_rate: float, L_eff: float) -> dict:
    """计算单列（给定降雨强度和有效路径长度）"""
    out = dict(vals)
    freq  = _fv(vals.get("工作频率",       "39"))
    elev  = _fv(vals.get("地面站仰角",     "10"))
    pol   = _fv(vals.get("极化倾角",       "45"))
    h_s   = _fv(vals.get("地面站海拔",     "0.0"))
    pres  = _fv(vals.get("大气压",         "1013.25"))
    temp  = _fv(vals.get("温度",           "288.15"))
    wv    = _fv(vals.get("水汽分压",       "10.0"))
    L_cld = _fv(vals.get("液态水柱含量",   "0.5"))
    N_T   = _fv(vals.get("电子总含量 TEC", "1e17"))
    sigma = _fv(vals.get("雨滴伪角",       "5"))

    # 雨衰
    from modules.rain_attenuation import calc_specific_attenuation
    g_arr, k, alpha = calc_specific_attenuation(
        freq, np.array([rain_rate]), elev, pol)
    gamma_R = float(g_arr[0])
    A_rain  = gamma_R * L_eff

    out["有效路径长度"] = f"{L_eff:.4f}"
    out["R"]            = f"{rain_rate:.2f}"
    out["γ_R"]          = f"{gamma_R:.4f}"
    out["链路雨衰"]     = f"{A_rain:.4f}"

    # 大气气体衰减
    ar = compute_atm_attenuation(
        freq_ghz=freq, elevation_deg=elev,
        pressure_hpa=pres, temperature_k=temp,
        water_vapor_hpa=wv, station_alt_km=h_s)
    out["O₂ 比衰减"]      = f"{ar.gamma_o2:.4f}"
    out["H₂O 比衰减"]     = f"{ar.gamma_h2o:.4f}"
    out["总气体比衰减"]   = f"{ar.gamma_total:.4f}"
    out["O₂ 路径衰减"]    = f"{ar.A_o2:.4f}"
    out["H₂O 路径衰减"]   = f"{ar.A_h2o:.4f}"
    out["总大气气体衰减"] = f"{ar.A_total:.4f}"

    # 云雾衰减
    cr = compute_cloud_attenuation(freq_ghz=freq, L_kg_m2=L_cld, elevation_deg=elev)
    out["云雾比衰减系数"] = f"{cr.KL:.4f}"
    out["云雾路径衰减"]   = f"{cr.A_cloud:.4f}"

    A_total = A_rain + ar.A_total + cr.A_cloud
    out["雨衰+大气+云雾"] = f"{A_total:.4f}"

    # 电离层群时延
    _, t_ns = calc_ionospheric_group_delay(freq, N_T)
    out["群时延"] = f"{t_ns:.4f}"

    # 雨致 XPD
    xpd = calc_xpd_rain(A_rain, freq, elev, pol, sigma)
    out["雨致 XPD"] = f"{xpd['XPD_rain']:.4f}"

    return out


def calc_city_columns(vals: dict, stats) -> list[dict]:
    """城市模式：利用 stats 生成 N_PROB 列结果"""
    freq = _fv(vals.get("工作频率", "39"))
    elev = _fv(vals.get("地面站仰角", "10"))
    h_s  = _fv(vals.get("地面站海拔", "0.0"))

    # L_eff 依赖 gamma_R(R001) 和频率（P.618-14 步骤6-8）
    from modules.rain_attenuation import calc_specific_attenuation as _csa
    g_arr0, _, _ = _csa(freq, np.array([stats.R001]), elev, 45.0)
    gamma_R_001 = float(g_arr0[0])
    L_eff = calc_effective_path_length(
        stats.rain_height_km, h_s, elev, stats.R001, stats.lat,
        freq_ghz=freq, gamma_R=gamma_R_001)

    results = []
    for pi in range(N_PROB):
        r = _calc_single(vals, stats.R_p_vals[pi], L_eff)
        # 城市模式 R_p 标签用 "R_p" 键覆盖 "R"（显示推导值）
        r["R"] = f"{stats.R_p_vals[pi]:.2f}"
        results.append(r)
    return results


def calc_custom_column(vals: dict, rain_rate: float,
                       rain_height_km: float, lat: float, R001: float) -> dict:
    """自定义模式：单场景列计算"""
    elev = _fv(vals.get("地面站仰角", "10"))
    h_s  = _fv(vals.get("地面站海拔", "0.0"))

    # 自定义模式：有效路径用简化公式（无 R001 时）或完整公式
    if R001 > 0:
        from modules.rain_attenuation import calc_specific_attenuation as _csa2
        freq2 = _fv(vals.get("工作频率", "39"))
        pol2  = _fv(vals.get("极化倾角", "45"))
        g2, _, _ = _csa2(freq2, np.array([R001]), elev, pol2)
        L_eff = calc_effective_path_length(
            rain_height_km, h_s, elev, R001, lat,
            freq_ghz=freq2, gamma_R=float(g2[0]))
    else:
        theta = math.radians(max(elev, 5.0))
        delta_h = max(rain_height_km - h_s, 0.0)
        L_eff = delta_h / math.sin(theta)

    return _calc_single(vals, rain_rate, L_eff)


# ══════════════════════════════════════════════════════════
#  编辑委托
# ══════════════════════════════════════════════════════════

class _EditDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        ed = QLineEdit(parent)
        ed.setStyleSheet(
            "QLineEdit{background:#FFFFFF;color:#111111;"
            "border:2px solid #2E6B8A;border-radius:2px;"
            "padding:1px 4px;font-size:9pt;"
            "selection-background-color:#BEDAF7;}")
        return ed
    def setEditorData(self, ed, index):
        ed.setText(index.data(Qt.ItemDataRole.EditRole) or "")
        ed.selectAll()
    def setModelData(self, ed, model, index):
        model.setData(index, ed.text(), Qt.ItemDataRole.EditRole)


# ══════════════════════════════════════════════════════════
#  雨衰图像对话框（仅城市模式）
# ══════════════════════════════════════════════════════════

class RainPlotDialog(QDialog):
    """
    城市模式雨衰统计图——横轴：降雨强度 R (mm/h)，纵轴：链路雨衰 A (dB)
    100 个等间距 R 点，背景标注 GB/T 28592-2012 降水等级色带。
    5 个超越概率对应的 (R_p, A_p) 散点叠加标注。
    """

    # GB/T 28592-2012 雨强等级（12h 累计量 ÷ 12，单位 mm/h）
    # 边界：0.42 / 1.25 / 2.5 / 5.83 / 11.67
    _RAIN_GRADES = [
        (0,       0.42,  (0.90, 0.95, 1.00), "微量"),
        (0.42,    1.25,  (0.82, 0.94, 0.82), "小雨"),
        (1.25,    2.50,  (0.60, 0.85, 0.60), "中雨"),
        (2.50,    5.83,  (0.35, 0.70, 0.35), "大雨"),
        (5.83,   11.67,  (0.20, 0.55, 0.85), "暴雨"),
        (11.67,  23.33,  (0.10, 0.35, 0.75), "大暴雨"),
        (23.33, 999.0,   (0.05, 0.15, 0.60), "特大暴雨"),
    ]

    def __init__(self, stats, parent=None):
        super().__init__(parent)
        self.setWindowTitle("雨衰统计图")
        self.setMinimumSize(820, 560)
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint)
        self._stats = stats
        self._build_ui()
        self._plot()

    def _build_ui(self):
        lv = QVBoxLayout(self)
        lv.setContentsMargins(8, 8, 8, 8)
        lv.setSpacing(6)
        self.canvas = FigureCanvas(Figure(figsize=(10, 5.5), dpi=120))
        self.canvas.figure.patch.set_facecolor("#F8F8F8")
        self.canvas.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        lv.addWidget(self.canvas)
        bhl = QHBoxLayout(); bhl.addStretch()
        btn = QPushButton("保存图像")
        btn.setFixedHeight(28)
        btn.setStyleSheet(
            "QPushButton{background:#2E6B8A;color:#FFF;border:none;"
            "border-radius:4px;padding:0 14px;font-size:9pt;}"
            "QPushButton:hover{background:#1E4D66;}")
        btn.clicked.connect(self._save)
        bhl.addWidget(btn)
        lv.addLayout(bhl)

    def _plot(self):
        from modules.rain_attenuation import calc_specific_attenuation, calc_effective_path_length
        from modules.cross_polarization import calc_xpd_rain
        import numpy as np

        s   = self._stats
        fig = self.canvas.figure
        fig.clf()
        fig.set_constrained_layout(True)
        ax_a = fig.add_subplot(111)
        ax_x = ax_a.twinx()
        ax_a.set_facecolor("#FAFCFF")
        fig.patch.set_facecolor("#F4F6F8")

        max_R = max(max(s.R_p_vals) * 1.4, 50.0)
        R_arr = np.linspace(0.1, max_R, 200)
        A_curve   = np.zeros_like(R_arr)
        XPD_curve = np.zeros_like(R_arr)

        for i, R in enumerate(R_arr):
            g_arr, _, _ = calc_specific_attenuation(
                s.freq_ghz, np.array([R]), s.elevation_deg, 45.0)
            gamma_i = float(g_arr[0])
            L = calc_effective_path_length(
                s.rain_height_km, 0.0, s.elevation_deg, s.freq_ghz, np.array([R]))
            A_i = gamma_i * float(L[0]) if hasattr(L, "__len__") else gamma_i * float(L)
            A_curve[i] = A_i
            xpd_res = calc_xpd_rain(A_i, s.freq_ghz, s.elevation_deg,
                                     tau_deg=45.0, sigma_deg=5.0)
            XPD_curve[i] = xpd_res["XPD_rain"]

        l1, = ax_a.plot(R_arr, A_curve, color="#CC2200", lw=2.0, zorder=3,
                        label="A (dB)")
        ax_a.axvline(s.R001, color="#999999", lw=1.0, ls="--", zorder=2)
        l2, = ax_x.plot(R_arr, XPD_curve, color="#0055CC", lw=2.0,
                        ls="-.", zorder=3, label="XPD (dB)")

        ax_a.set_xlim(0, max_R)
        y_top = max(A_curve) * 1.18
        ax_a.set_ylim(0, y_top)
        ax_a.set_xlabel("R (mm/h)", fontsize=11)
        ax_a.set_ylabel("A (dB)", fontsize=11, color="#CC2200")
        ax_a.tick_params(axis="y", labelcolor="#CC2200", labelsize=9.5)
        ax_a.tick_params(axis="x", labelsize=9.5)

        xpd_valid = XPD_curve[np.isfinite(XPD_curve)]
        xpd_lo = max(0, float(np.min(xpd_valid)) - 3) if len(xpd_valid) else 0
        xpd_hi = min(80, float(np.max(xpd_valid)) + 5) if len(xpd_valid) else 40
        ax_x.set_ylim(xpd_lo, xpd_hi)
        ax_x.set_ylabel("XPD (dB)", fontsize=11, color="#0055CC")
        ax_x.tick_params(axis="y", labelcolor="#0055CC", labelsize=9.5)
        ax_x.spines["right"].set_color("#0055CC")
        ax_a.spines["left"].set_color("#CC2200")
        for sp in ["top", "bottom"]:
            ax_a.spines[sp].set_color("#CCCCCC")
            ax_x.spines[sp].set_visible(False)
        ax_x.spines["left"].set_visible(False)
        ax_a.grid(True, color="#E4E8EE", lw=0.6, zorder=1)

        handles = [l1, l2,
                   plt.Line2D([0],[0], color="#999", lw=1, ls="--",
                               label=f"R₀.₀₁={s.R001:.1f} mm/h")]
        ax_a.legend(handles=handles, fontsize=9, loc="upper left",
                    framealpha=0.93, edgecolor="#CCC")
        ax_a.set_title(
            f"{s.city}  f={s.freq_ghz} GHz  θ={s.elevation_deg}°  "
            f"H_R={s.rain_height_km} km  R_0.01={s.R001:.1f} mm/h",
            fontsize=10, pad=8)
        self.canvas.draw()
    def _save(self):
        city = self._stats.city or "rain"
        path, _ = QFileDialog.getSaveFileName(
            self, "保存图像", f"rain_{city}.png",
            "PNG (*.png);;PDF (*.pdf);;SVG (*.svg)")
        if path:
            self.canvas.figure.savefig(path, dpi=150, bbox_inches="tight")
            QMessageBox.information(self, "保存成功", f"已保存：\n{path}")
# ══════════════════════════════════════════════════════════
#  主 Widget
# ══════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════
#  群时延图对话框
# ══════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════
#  雨衰频域图对话框
#  横轴：频率 (GHz)，纵轴：链路雨衰 A (dB)
#  输入：中心频率、带宽、降雨强度 R、仰角、雨顶高度
# ══════════════════════════════════════════════════════════

class RainFreqDialog(QDialog):
    """
    雨衰随频率变化曲线。
    对每个频率点计算：gamma_R(f, R) × L_eff(f, R)。
    用户可指定：中心频率 fc (GHz)、带宽 BW (MHz)、
               降雨强度 R (mm/h)、仰角 θ (°)、雨顶高度 H_R (km)。
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("雨衰频域图")
        self.setMinimumSize(680, 530)
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint)
        self._build_ui()

    def _build_ui(self):
        lv = QVBoxLayout(self)
        lv.setContentsMargins(8, 8, 8, 8)
        lv.setSpacing(6)

        _es = ("QLineEdit{background:#FFF;border:1px solid #C0C8D8;"
               "border-radius:3px;padding:2px 6px;font-size:9pt;}")
        _lb = "font-size:9pt;"

        # ── 参数行 ────────────────────────────────────────
        row1 = QHBoxLayout(); row1.setSpacing(10)
        row1.addWidget(QLabel("中心频率 (GHz):", styleSheet=_lb))
        self.e_fc = QLineEdit("39"); self.e_fc.setFixedWidth(65); self.e_fc.setStyleSheet(_es)
        row1.addWidget(self.e_fc)
        row1.addWidget(QLabel("带宽 (MHz):", styleSheet=_lb))
        self.e_bw = QLineEdit("2000"); self.e_bw.setFixedWidth(65); self.e_bw.setStyleSheet(_es)
        row1.addWidget(self.e_bw)
        row1.addWidget(QLabel("降雨强度 R (mm/h):", styleSheet=_lb))
        self.e_R = QLineEdit("30"); self.e_R.setFixedWidth(55); self.e_R.setStyleSheet(_es)
        row1.addWidget(self.e_R)
        row1.addStretch()
        lv.addLayout(row1)

        row2 = QHBoxLayout(); row2.setSpacing(10)
        row2.addWidget(QLabel("仰角 θ (°):", styleSheet=_lb))
        self.e_el = QLineEdit("10"); self.e_el.setFixedWidth(50); self.e_el.setStyleSheet(_es)
        row2.addWidget(self.e_el)
        row2.addWidget(QLabel("雨顶高度 H_R (km):", styleSheet=_lb))
        self.e_hr = QLineEdit("4.0"); self.e_hr.setFixedWidth(55); self.e_hr.setStyleSheet(_es)
        row2.addWidget(self.e_hr)
        row2.addWidget(QLabel("极化倾角 τ (°):", styleSheet=_lb))
        self.e_tau = QLineEdit("45"); self.e_tau.setFixedWidth(50); self.e_tau.setStyleSheet(_es)
        row2.addWidget(self.e_tau)
        btn_draw = QPushButton("绘制")
        btn_draw.setFixedHeight(26)
        btn_draw.setStyleSheet("QPushButton{background:#8A5B2E;color:#FFF;border:none;"
                                "border-radius:3px;padding:0 12px;font-size:9pt;}"
                                "QPushButton:hover{background:#5C3D1E;}")
        btn_draw.clicked.connect(self._plot)
        row2.addWidget(btn_draw)
        row2.addStretch()
        btn_save = QPushButton("保存图像")
        btn_save.setFixedHeight(26)
        btn_save.setStyleSheet("QPushButton{background:#2E6B8A;color:#FFF;border:none;"
                                "border-radius:3px;padding:0 12px;font-size:9pt;}"
                                "QPushButton:hover{background:#1E4D66;}")
        btn_save.clicked.connect(self._save)
        row2.addWidget(btn_save)
        lv.addLayout(row2)

        self.canvas = FigureCanvas(Figure(figsize=(8, 5), dpi=110))
        self.canvas.figure.patch.set_facecolor("#F4F6F8")
        self.canvas.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        lv.addWidget(self.canvas)

    def _plot(self):
        import numpy as np
        from modules.rain_attenuation import (
            calc_specific_attenuation,
            calc_effective_path_length,
        )
        try:
            fc  = float(self.e_fc.text())
            bw  = float(self.e_bw.text()) / 1e3   # MHz -> GHz
            R   = float(self.e_R.text())
            el  = float(self.e_el.text())
            hr  = float(self.e_hr.text())
            tau = float(self.e_tau.text())
        except ValueError:
            return
        if fc <= 0 or bw <= 0 or R <= 0:
            return

        f_lo  = max(fc - bw/2, 1.0)
        f_hi  = fc + bw/2
        f_arr = np.linspace(f_lo, f_hi, 300)
        A_arr = np.zeros_like(f_arr)

        for i, f in enumerate(f_arr):
            g_arr, _, _ = calc_specific_attenuation(
                f, np.array([R]), el, tau)
            gamma_i = float(g_arr[0])
            # 用 R 作为 R001，调用用户版本签名（freq_ghz, rain_rate 为 numpy数组）
            L = calc_effective_path_length(
                hr, 0.0, el, f, np.array([R]))
            L_val = float(L[0]) if hasattr(L, "__len__") else float(L)
            A_arr[i] = gamma_i * L_val

        fig = self.canvas.figure
        fig.clf()
        fig.set_constrained_layout(True)
        ax = fig.add_subplot(111)
        ax.set_facecolor("#FAFCFF")
        fig.patch.set_facecolor("#F4F6F8")

        ax.plot(f_arr, A_arr, color="#8A5B2E", lw=2.0)
        ax.axvline(fc, color="#888", lw=1.0, ls="--",
                   label=f"fc={fc} GHz")
        ax.axvline(f_lo, color="#BBBBBB", lw=0.8, ls=":",
                   label=f"f_lo={f_lo:.3f} GHz")
        ax.axvline(f_hi, color="#BBBBBB", lw=0.8, ls=":",
                   label=f"f_hi={f_hi:.3f} GHz")

        # 标注中心频率处的雨衰值
        g_fc, _, _ = calc_specific_attenuation(fc, np.array([R]), el, tau)
        L_fc = calc_effective_path_length(hr, 0.0, el, fc, np.array([R]))
        A_fc = float(g_fc[0]) * (float(L_fc[0]) if hasattr(L_fc, "__len__") else float(L_fc))
        ax.annotate(
            f"A(fc)={A_fc:.3f} dB",
            xy=(fc, A_fc), xytext=(12, 10),
            textcoords="offset points",
            fontsize=9, color="#8A5B2E",
            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#8A5B2E", alpha=0.9, lw=0.8),
            arrowprops=dict(arrowstyle="-", color="#8A5B2E", lw=0.8))

        # 带内雨衰变化量
        delta_A = float(np.max(A_arr) - np.min(A_arr))
        ax.set_xlabel("频率 (GHz)", fontsize=11)
        ax.set_ylabel("链路雨衰 A (dB)", fontsize=11)
        ax.set_title(
            f"雨衰随频率变化  R={R} mm/h  θ={el}°  H_R={hr} km  τ={tau}°\n"
            f"fc={fc} GHz  BW={self.e_bw.text()} MHz  ΔA={delta_A:.3f} dB",
            fontsize=10, pad=8)
        ax.legend(fontsize=8.5, framealpha=0.93, edgecolor="#CCC")
        ax.grid(True, color="#E4E8EE", lw=0.6)
        for sp in ax.spines.values():
            sp.set_color("#CCCCCC"); sp.set_linewidth(0.8)
        ax.tick_params(labelsize=9.5)
        self.canvas.draw()

    def _save(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存图像", "rain_freq.png",
            "PNG (*.png);;PDF (*.pdf);;SVG (*.svg)")
        if path:
            self.canvas.figure.savefig(path, dpi=150, bbox_inches="tight")
            QMessageBox.information(self, "保存成功", f"已保存：\n{path}")


class GroupDelayDialog(QDialog):
    """
    电离层群时延随频率变化曲线。
    t(f) = 1.345e-7 * N_T / f^2  (s)，换算为 ns。
    横轴：频率 (GHz)，由中心频率和带宽确定范围。
    """

    def __init__(self, N_T=1e17, parent=None):
        super().__init__(parent)
        self.setWindowTitle("电离层群时延")
        self.setMinimumSize(640, 500)
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint |
            Qt.WindowType.WindowCloseButtonHint)
        self._N_T = N_T
        self._build_ui()
        self._plot()

    def _build_ui(self):
        lv = QVBoxLayout(self)
        lv.setContentsMargins(8, 8, 8, 8)
        lv.setSpacing(6)

        # 控制栏
        ctrl = QHBoxLayout(); ctrl.setSpacing(8)
        ctrl.addWidget(QLabel("中心频率 (GHz):", styleSheet="font-size:9pt;"))
        self.e_fc = QLineEdit("39"); self.e_fc.setFixedWidth(70)
        self.e_fc.setStyleSheet("QLineEdit{background:#FFF;border:1px solid #C0C8D8;border-radius:3px;padding:2px 6px;font-size:9pt;}")
        ctrl.addWidget(self.e_fc)
        ctrl.addWidget(QLabel("带宽 (MHz):", styleSheet="font-size:9pt;"))
        self.e_bw = QLineEdit("500"); self.e_bw.setFixedWidth(70)
        self.e_bw.setStyleSheet(self.e_fc.styleSheet())
        ctrl.addWidget(self.e_bw)
        btn_draw = QPushButton("绘制")
        btn_draw.setFixedHeight(26)
        btn_draw.setStyleSheet("QPushButton{background:#2E6B8A;color:#FFF;border:none;border-radius:3px;padding:0 12px;font-size:9pt;}QPushButton:hover{background:#1E4D66;}")
        btn_draw.clicked.connect(self._plot)
        ctrl.addWidget(btn_draw)
        ctrl.addStretch()
        btn_save = QPushButton("保存图像")
        btn_save.setFixedHeight(26)
        btn_save.setStyleSheet("QPushButton{background:#6B4C2E;color:#FFF;border:none;border-radius:3px;padding:0 12px;font-size:9pt;}QPushButton:hover{background:#4A3420;}")
        btn_save.clicked.connect(self._save)
        ctrl.addWidget(btn_save)
        lv.addLayout(ctrl)

        self.canvas = FigureCanvas(Figure(figsize=(8, 5), dpi=110))
        self.canvas.figure.patch.set_facecolor("#F4F6F8")
        self.canvas.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        lv.addWidget(self.canvas)

    def _plot(self):
        import numpy as np
        from modules.ionosphere_effects import calc_ionospheric_group_delay
        try:
            fc  = float(self.e_fc.text())
            bw  = float(self.e_bw.text()) / 1e3   # MHz -> GHz
        except ValueError:
            return
        if fc <= 0 or bw <= 0:
            return

        f_lo  = max(fc - bw/2, 0.1)
        f_hi  = fc + bw/2
        f_arr = np.linspace(f_lo, f_hi, 300)
        t_arr = np.array([calc_ionospheric_group_delay(f, self._N_T)[1]
                          for f in f_arr])   # ns

        fig = self.canvas.figure
        fig.clf()
        fig.set_constrained_layout(True)
        ax  = fig.add_subplot(111)
        ax.set_facecolor("#FAFCFF")
        fig.patch.set_facecolor("#F4F6F8")

        ax.plot(f_arr, t_arr, color="#1D9E75", lw=2.0)
        ax.axvline(fc, color="#888", lw=1.0, ls="--",
                   label=f"中心频率 {fc} GHz")
        ax.axvline(f_lo, color="#BBBBBB", lw=0.8, ls=":", label=f"f_lo={f_lo:.3f} GHz")
        ax.axvline(f_hi, color="#BBBBBB", lw=0.8, ls=":", label=f"f_hi={f_hi:.3f} GHz")

        # 标注中心频率处的群时延值
        t_fc = calc_ionospheric_group_delay(fc, self._N_T)[1]
        ax.annotate(
            f"t(fc)={t_fc:.4f} ns",
            xy=(fc, t_fc), xytext=(12, -18),
            textcoords="offset points",
            fontsize=9, color="#1D9E75",
            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#1D9E75", alpha=0.9, lw=0.8),
            arrowprops=dict(arrowstyle="-", color="#1D9E75", lw=0.8))

        # 差分群时延（带内时延差）
        delta_t = float(np.max(t_arr) - np.min(t_arr))
        ax.set_xlabel("频率 (GHz)", fontsize=11)
        ax.set_ylabel("群时延 (ns)", fontsize=11)
        ax.set_title(
            f"电离层群时延  TEC={self._N_T:.1e} el/m²\n"
            f"fc={fc} GHz  BW={self.e_bw.text()} MHz  Δt={delta_t:.4f} ns",
            fontsize=10, pad=8)
        ax.legend(fontsize=8.5, framealpha=0.93, edgecolor="#CCC")
        ax.grid(True, color="#E4E8EE", lw=0.6)
        for sp in ax.spines.values():
            sp.set_color("#CCCCCC"); sp.set_linewidth(0.8)
        ax.tick_params(labelsize=9.5)

        self.canvas.draw()

    def _save(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存图像", "group_delay.png",
            "PNG (*.png);;PDF (*.pdf);;SVG (*.svg)")
        if path:
            self.canvas.figure.savefig(path, dpi=150, bbox_inches="tight")
            QMessageBox.information(self, "保存成功", f"已保存：\n{path}")


class ChannelTableWidget(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self._updating   = False
        self._last_stats = None
        self._R001       = 42.0      # 城市模式 R_0.01 缓存
        self._city_mode  = False     # False=自定义, True=城市
        self._n_custom   = 3         # 自定义模式场景列数
        self._build_ui()

    # ── 界面构建 ──────────────────────────────────────────

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(6, 6, 6, 6)
        root.setSpacing(6)

        # 工具栏
        tb = QHBoxLayout(); tb.setSpacing(6)

        tb.addWidget(QLabel("城市：", styleSheet="font-size:9pt;color:#444;"))
        self.cb_city = QComboBox()
        self.cb_city.setFixedWidth(110)
        self.cb_city.setStyleSheet(
            "QComboBox{background:#FFF;border:1px solid #C0C8D8;"
            "border-radius:3px;padding:2px 6px;font-size:9pt;color:#111;}"
            "QComboBox QAbstractItemView{background:#FFF;font-size:9pt;"
            "color:#111;selection-background-color:#D0E8F8;}")
        self.cb_city.addItem("自定义")
        self.cb_city.addItems(CITY_NAMES)
        self.cb_city.currentTextChanged.connect(self._on_city_changed)
        tb.addWidget(self.cb_city)

        self.lbl_r001 = QLabel("R₀.₀₁=— mm/h")
        self.lbl_r001.setStyleSheet("font-size:8pt;color:#666;")
        tb.addWidget(self.lbl_r001)

        self.lbl_src = QLabel(data_source_info())
        self.lbl_src.setStyleSheet("font-size:8pt;color:#999;")
        tb.addWidget(self.lbl_src)
        tb.addStretch()

        # 自定义模式按钮（城市模式时隐藏）
        self.btn_add = QPushButton("＋ 添加场景")
        self.btn_add.setFixedHeight(28)
        self.btn_add.setStyleSheet(self._bstyle("#2E6B8A"))
        self.btn_add.clicked.connect(self._add_custom_col)
        tb.addWidget(self.btn_add)

        self.btn_del = QPushButton("删除选中列")
        self.btn_del.setFixedHeight(28)
        self.btn_del.setStyleSheet(self._bstyle("#888888"))
        self.btn_del.clicked.connect(self._del_custom_col)
        tb.addWidget(self.btn_del)

        # 城市模式按钮（自定义模式时隐藏）
        self.btn_plot = QPushButton("雨衰与去极化图像")
        self.btn_plot.setFixedHeight(28)
        self.btn_plot.setStyleSheet(self._bstyle("#1D6B42"))
        self.btn_plot.clicked.connect(self._show_rain_plot)
        tb.addWidget(self.btn_plot)

        self.btn_delay = QPushButton("群时延图像")
        self.btn_delay.setFixedHeight(28)
        self.btn_delay.setStyleSheet(self._bstyle("#5B3FA0"))
        self.btn_delay.clicked.connect(self._show_group_delay)
        tb.addWidget(self.btn_delay)

        self.btn_rain_freq = QPushButton("雨衰频域图像")
        self.btn_rain_freq.setFixedHeight(28)
        self.btn_rain_freq.setStyleSheet(self._bstyle("#8A5B2E"))
        self.btn_rain_freq.clicked.connect(self._show_rain_freq)
        tb.addWidget(self.btn_rain_freq)

        self.btn_excel = QPushButton("导出 Excel")
        self.btn_excel.setFixedHeight(28)
        self.btn_excel.setStyleSheet(self._bstyle("#6B4C2E"))
        self.btn_excel.clicked.connect(self._export_excel)
        tb.addWidget(self.btn_excel)

        root.addLayout(tb)

        # 表格
        self.table = QTableWidget()
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background:#FFFFFF; border:1px solid #C8D8EC;
                gridline-color:#D8E4F0; font-size:9pt; color:#1A1A1A;
            }}
            QTableWidget::item {{ padding:2px 5px; }}
            QTableWidget::item:selected {{ background:#C5DCF5; color:#1A1A1A; }}
            QHeaderView::section {{
                background:{_C_HDR_BG}; color:#FFFFFF;
                font-size:9pt; font-weight:600;
                padding:4px 5px; border:none;
                border-right:1px solid #4A7EA0;
                border-bottom:1px solid #4A7EA0;
            }}
        """)
        self.table.setAlternatingRowColors(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked |
            QAbstractItemView.EditTrigger.AnyKeyPressed)
        self.table.setItemDelegate(_EditDelegate(self.table))
        self.table.itemChanged.connect(self._on_changed)
        root.addWidget(self.table, stretch=1)

        self.status = QLabel("就绪")
        self.status.setStyleSheet("font-size:9pt;color:#888;")
        root.addWidget(self.status)

        # 初始为自定义模式
        self._rebuild_table()

    # ── 按钮状态切换 ──────────────────────────────────────

    def _update_buttons(self):
        self.btn_add.setVisible(not self._city_mode)
        self.btn_del.setVisible(not self._city_mode)
        self.btn_plot.setVisible(self._city_mode)
        self.btn_delay.setVisible(True)      # 两种模式均可用
        self.btn_rain_freq.setVisible(True)   # 两种模式均可用

    # ══════════════════════════════════════════════════════
    #  表格构建
    # ══════════════════════════════════════════════════════

    def _rebuild_table(self):
        """完整重建表格（切换模式时调用）"""
        self._updating = True
        n_cols_data = N_PROB if self._city_mode else self._n_custom
        n_cols = 3 + n_cols_data

        self.table.setRowCount(len(ROWS))
        self.table.setColumnCount(n_cols)

        if self._city_mode:
            col_labels = PROB_LABELS
        else:
            col_labels = [f"场景 {i+1}" for i in range(self._n_custom)]
        self.table.setHorizontalHeaderLabels(["参数","符号","单位"] + col_labels)

        self.table.setColumnWidth(0, 165)
        self.table.setColumnWidth(1, 55)
        self.table.setColumnWidth(2, 72)
        for ci in range(3, n_cols):
            self.table.setColumnWidth(ci, 105)

        hdr = self.table.horizontalHeader()
        for ci in range(n_cols):
            hdr.setSectionResizeMode(ci, QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(False)

        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            self.table.setRowHeight(ri, 22)
            self._set_meta(ri, 0, name, rtype)
            self._set_meta(ri, 1, sym,  rtype)
            self._set_meta(ri, 2, unit, rtype)
            for di in range(n_cols_data):
                self._init_cell(ri, 3 + di)

        self._updating = False
        self._update_buttons()
        self._calc_all()

    def _set_meta(self, ri, ci, text, rtype):
        item = QTableWidgetItem(text)
        item.setFlags(Qt.ItemFlag.ItemIsEnabled)
        if rtype == "section":
            item.setBackground(QBrush(_C_SECTION))
            f = item.font(); f.setBold(True); item.setFont(f)
            item.setForeground(QBrush(QColor("#1A4A80")))
        else:
            item.setBackground(QBrush(QColor("#F5F8FC")))
            item.setForeground(QBrush(QColor("#444444")))
        self.table.setItem(ri, ci, item)

    def _init_cell(self, ri, ci):
        """初始化一个数据格子"""
        name, sym, unit, rtype = ROWS[ri]

        if rtype == "section":
            item = QTableWidgetItem("")
            item.setFlags(Qt.ItemFlag.NoItemFlags)
            item.setBackground(QBrush(_C_SECTION))
            self.table.setItem(ri, ci, item)
            return

        item = QTableWidgetItem(DEFAULTS.get(name, ""))

        if rtype == "coord":
            # 只读，浅灰背景，城市模式填入
            item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                          Qt.ItemFlag.ItemIsSelectable)
            item.setBackground(QBrush(QColor("#F5F8FC")))
            item.setForeground(QBrush(QColor("#555555")))

        elif rtype == "rain_in":
            if self._city_mode:
                # 城市模式：降雨强度是计算结果，不可编辑
                item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                              Qt.ItemFlag.ItemIsSelectable)
                item.setBackground(QBrush(_C_CALC))
            else:
                # 自定义模式：可编辑，各列独立
                item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                              Qt.ItemFlag.ItemIsSelectable |
                              Qt.ItemFlag.ItemIsEditable)
                item.setBackground(QBrush(_C_INPUT))

        elif rtype == "input":
            # 城市模式：只有第一列可编辑（其余同步）
            # 自定义模式：所有列各自独立可编辑
            if self._city_mode and ci > 3:
                item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                              Qt.ItemFlag.ItemIsSelectable)
            else:
                item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                              Qt.ItemFlag.ItemIsSelectable |
                              Qt.ItemFlag.ItemIsEditable)
            item.setBackground(QBrush(_C_INPUT))

        else:  # calc
            item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                          Qt.ItemFlag.ItemIsSelectable)
            item.setBackground(QBrush(_C_CALC))

        self.table.setItem(ri, ci, item)

    # ══════════════════════════════════════════════════════
    #  自定义模式增/删列
    # ══════════════════════════════════════════════════════

    def _add_custom_col(self):
        self._n_custom += 1
        ci = self.table.columnCount()
        self.table.insertColumn(ci)
        self.table.setHorizontalHeaderItem(
            ci, QTableWidgetItem(f"场景 {self._n_custom}"))
        self.table.setColumnWidth(ci, 105)
        self.table.horizontalHeader().setSectionResizeMode(
            ci, QHeaderView.ResizeMode.Interactive)
        self._updating = True
        for ri in range(len(ROWS)):
            self._init_cell(ri, ci)
        self._updating = False
        self._calc_col_custom(ci)

    def _del_custom_col(self):
        sel = sorted(
            {idx.column() for idx in self.table.selectedIndexes()
             if idx.column() >= 3}, reverse=True)
        if not sel:
            QMessageBox.information(self, "提示", "请先选中要删除的场景列")
            return
        if self._n_custom - len(sel) < 1:
            QMessageBox.information(self, "提示", "至少保留一列")
            return
        for c in sel:
            self.table.removeColumn(c)
            self._n_custom -= 1
        for idx2, ci2 in enumerate(range(3, self.table.columnCount())):
            self.table.setHorizontalHeaderItem(
                ci2, QTableWidgetItem(f"场景 {idx2+1}"))

    # ══════════════════════════════════════════════════════
    #  城市选择
    # ══════════════════════════════════════════════════════

    def _on_city_changed(self, city_name: str):
        if self._updating:
            return

        is_custom = (city_name == "自定义")
        mode_changed = (is_custom == self._city_mode)  # 模式是否切换

        if is_custom:
            if self._city_mode:           # 从城市模式切到自定义
                self._city_mode = False
                self._R001 = 0.0
                self.lbl_r001.setText("R₀.₀₁=— mm/h")
                self._n_custom = 3
                self._rebuild_table()
            # 已经是自定义则无需重建
        else:
            params = get_city_rain_params(city_name)
            self._R001 = params["R001_mmh"]
            self.lbl_r001.setText(f"R₀.₀₁={self._R001:.1f} mm/h")

            if not self._city_mode:      # 从自定义切到城市模式
                self._city_mode = True
                self._rebuild_table()    # rebuild 后再填坐标

            # 填入 coord 行和雨顶高度
            self._updating = True
            for ri, (name, sym, unit, rtype) in enumerate(ROWS):
                val = None
                if rtype == "input" and name == "雨顶高度":
                    val = str(params["rain_height_km"])
                elif rtype == "coord":
                    if name == "地面站纬度":
                        val = str(round(params["lat"], 4))
                    elif name == "地面站经度":
                        val = str(round(params["lon"], 4))
                if val is not None:
                    for pi in range(N_PROB):
                        it = self.table.item(ri, 3 + pi)
                        if it:
                            it.setText(val)
            self._updating = False
            self._calc_all()
            self.status.setText(
                f"已加载 {city_name}：H_R={params['rain_height_km']} km  "
                f"R₀.₀₁={self._R001:.1f} mm/h  "
                f"lat={params['lat']:.2f}°  lon={params['lon']:.2f}°")

    # ══════════════════════════════════════════════════════
    #  编辑响应
    # ══════════════════════════════════════════════════════

    def _on_changed(self, item):
        if self._updating:
            return
        ri = item.row()
        ci = item.column()
        if ci < 3:
            return

        name, sym, unit, rtype = ROWS[ri]

        if self._city_mode:
            # 城市模式：input 行第一列编辑时同步其他列
            if ci == 3 and rtype == "input":
                val = item.text()
                self._updating = True
                for pi in range(1, N_PROB):
                    it = self.table.item(ri, 3 + pi)
                    if it and not (it.flags() & Qt.ItemFlag.ItemIsEditable == Qt.ItemFlag.NoItemFlags):
                        it.setText(val)
                self._updating = False

        self._calc_all()

    # ══════════════════════════════════════════════════════
    #  计算
    # ══════════════════════════════════════════════════════

    def _read_shared_vals(self) -> dict:
        """读取第一列（城市模式）或各列共用的 input 值"""
        vals = {}
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if rtype == "input":
                it = self.table.item(ri, 3)
                vals[name] = it.text() if it else DEFAULTS.get(name, "")
        return vals

    def _read_coord(self, row_name: str, default: float) -> float:
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if name == row_name and rtype == "coord":
                it = self.table.item(ri, 3)
                txt = (it.text() if it else "").strip()
                return _fv(txt, default) if txt else default
        return default

    def _calc_all(self):
        if self._city_mode:
            self._calc_all_city()
        else:
            self._calc_all_custom()

    def _calc_all_city(self):
        vals   = self._read_shared_vals()
        freq   = _fv(vals.get("工作频率",   "39"))
        elev   = _fv(vals.get("地面站仰角", "10"))
        pol    = _fv(vals.get("极化倾角",   "45"))
        h_rain = _fv(vals.get("雨顶高度",   "4.0"))
        lat    = self._read_coord("地面站纬度", 35.0)

        try:
            stats = compute_rain_statistics(
                freq_ghz=freq, elevation_deg=elev,
                polarization_deg=pol,
                rain_height_km=h_rain,
                R001=self._R001,
                lat=lat,
                station_alt_km=_fv(vals.get("地面站海拔","0.0")),
                city=self.cb_city.currentText(),
            )
            self._last_stats = stats
        except Exception as e:
            self.status.setText(f"雨衰统计计算错误：{e}")
            return

        try:
            all_results = calc_city_columns(vals, stats)
        except Exception as e:
            self.status.setText(f"计算错误：{e}")
            return

        self._updating = True
        for pi, result in enumerate(all_results):
            self._write_col(3 + pi, result)
        self._updating = False

        self.status.setText(
            f"城市模式 | f={freq}GHz  θ={elev}°  R₀.₀₁={self._R001:.1f}mm/h  "
            f"H_R={h_rain}km  | "
            f"A(0.01%)={all_results[1].get('链路雨衰','—')}dB  "
            f"A_total(0.01%)={all_results[1].get('雨衰+大气+云雾','—')}dB")

    def _calc_all_custom(self):
        n_cols_data = self.table.columnCount() - 3
        for di in range(n_cols_data):
            self._calc_col_custom(3 + di)

    def _calc_col_custom(self, ci: int):
        """计算自定义模式单列"""
        vals = {}
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if rtype == "input":
                # 自定义模式各列独立读取
                it = self.table.item(ri, ci)
                vals[name] = it.text() if it else DEFAULTS.get(name, "")

        # 读降雨强度（rain_in 行）
        rain_it = self.table.item(_ROW_R, ci)
        rain_rate = _fv(rain_it.text() if rain_it else "", 30.0)

        h_rain = _fv(vals.get("雨顶高度", "4.0"))
        h_s    = _fv(vals.get("地面站海拔","0.0"))
        lat    = self._read_coord("地面站纬度", 35.0)

        try:
            result = calc_custom_column(vals, rain_rate, h_rain, lat, self._R001)
        except Exception as e:
            self.status.setText(f"场景{ci-2} 计算错误：{e}")
            return

        self._updating = True
        self._write_col(ci, result)
        self._updating = False

    def _write_col(self, ci: int, result: dict):
        """将计算结果写入指定列"""
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            # 城市模式：calc 和 rain_in 行写入；自定义模式：只写 calc 行
            should_write = False
            if rtype == "calc":
                should_write = True
            elif rtype == "rain_in" and self._city_mode:
                should_write = True   # 城市模式降雨强度由计算填入

            if not should_write:
                continue

            it = self.table.item(ri, ci)
            if it is None:
                it = QTableWidgetItem()
                it.setFlags(Qt.ItemFlag.ItemIsEnabled |
                            Qt.ItemFlag.ItemIsSelectable)
                it.setBackground(QBrush(_C_CALC))
                self.table.setItem(ri, ci, it)

            val_str = result.get(name, "")
            it.setText(val_str)
            self._apply_color(it, ri, val_str)

    def _apply_color(self, it, ri, val_str):
        if ri == _ROW_TOTAL:
            try:
                fv = float(val_str)
                bg = QColor("#E3F2FD") if fv<=15 else QColor("#FCE4EC")
                fg = QColor("#0D47A1") if fv<=15 else QColor("#880E4F")
                it.setBackground(QBrush(bg)); it.setForeground(QBrush(fg))
                f = it.font(); f.setBold(True); it.setFont(f)
            except ValueError: pass
        elif ri in (_ROW_RAIN, _ROW_GAS, _ROW_CLOUD):
            try:
                fv = float(val_str)
                bg = QColor("#E8F5E9") if fv<=10 else QColor("#FFF3E0")
                fg = QColor("#1A6B35") if fv<=10 else QColor("#B05000")
                it.setBackground(QBrush(bg)); it.setForeground(QBrush(fg))
                f = it.font(); f.setBold(True); it.setFont(f)
            except ValueError: pass
        elif ri == _ROW_XPD_R:
            try:
                fv = float(val_str)
                if fv>=25:   bg,fg = QColor("#E8F5E9"),QColor("#1A6B35")
                elif fv>=10: bg,fg = QColor("#FFF3E0"),QColor("#B05000")
                else:        bg,fg = QColor("#FCE4EC"),QColor("#880E4F")
                it.setBackground(QBrush(bg)); it.setForeground(QBrush(fg))
                f = it.font(); f.setBold(True); it.setFont(f)
            except ValueError: pass

    # ══════════════════════════════════════════════════════
    #  输出图像（仅城市模式）
    # ══════════════════════════════════════════════════════

    def _show_rain_plot(self):
        if self._last_stats is None:
            QMessageBox.information(self, "提示", "请先选择城市")
            return
        dlg = RainPlotDialog(self._last_stats, parent=self)
        dlg.show()

    def _show_rain_freq(self):
        """雨衰频域图像——横轴频率，用户指定中心频率/带宽/降雨强度"""
        # 从表格读取工作频率作为中心频率默认值
        fc_default = "39"
        R_default  = "30"
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if name == "工作频率" and rtype == "input":
                it = self.table.item(ri, 3)
                if it and it.text().strip():
                    fc_default = it.text().strip()
                break
        # 自定义模式：读第一列的降雨强度
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if name == "R" and rtype == "rain_in":
                it = self.table.item(ri, 3)
                if it and it.text().strip():
                    R_default = it.text().strip()
                break
        dlg = RainFreqDialog(parent=self)
        dlg.e_fc.setText(fc_default)
        dlg.e_R.setText(R_default)
        dlg._plot()
        dlg.show()

    def _show_group_delay(self):
        N_T = 1e17   # 默认 TEC
        # 从表格读取用户输入的 TEC（如有）
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if name == "电子总含量 TEC" and rtype == "input":
                it = self.table.item(ri, 3)
                if it and it.text().strip():
                    try: N_T = float(it.text().strip())
                    except: pass
                break
        # 从表格读取工作频率作为中心频率默认值
        fc_default = "39"
        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            if name == "工作频率" and rtype == "input":
                it = self.table.item(ri, 3)
                if it and it.text().strip():
                    fc_default = it.text().strip()
                break
        dlg = GroupDelayDialog(N_T=N_T, parent=self)
        dlg.e_fc.setText(fc_default)
        dlg._plot()
        dlg.show()

    # ══════════════════════════════════════════════════════
    #  导出 Excel
    # ══════════════════════════════════════════════════════

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel", "信道建模.xlsx", "Excel (*.xlsx)")
        if not path: return
        try:
            self._write_excel(path)
            QMessageBox.information(self, "导出成功", f"已保存：\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", str(e))

    def _write_excel(self, path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook(); ws = wb.active; ws.title = "信道建模"
        thin = Side(style="thin", color="C8D8EC")
        brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr  = Alignment(horizontal="center", vertical="center")
        lft  = Alignment(horizontal="left",   vertical="center")
        def fill(h): return PatternFill("solid", start_color=h, end_color=h)

        # 说明行
        mode_str = (f"城市模式  城市={self.cb_city.currentText()}"
                    f"  R₀.₀₁={self._R001:.2f} mm/h"
                    if self._city_mode else "自定义模式")
        ws.append([mode_str])
        ws.row_dimensions[1].height = 18
        ws['A1'].font = Font(size=9, italic=True, color="666666")

        n_data = N_PROB if self._city_mode else self._n_custom
        col_labels = PROB_LABELS if self._city_mode else \
            [self.table.horizontalHeaderItem(3+i).text() for i in range(n_data)]

        hdrs = ["参数","符号","单位"] + col_labels
        ws.row_dimensions[2].height = 24
        for ci_x, h in enumerate(hdrs, 1):
            c = ws.cell(2, ci_x, h)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = fill("2E6B8A"); c.alignment = ctr; c.border = brd

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 7
        ws.column_dimensions["C"].width = 14
        for i in range(n_data):
            ws.column_dimensions[get_column_letter(4+i)].width = 12

        for ri, (name, sym, unit, rtype) in enumerate(ROWS):
            er = ri + 3
            ws.row_dimensions[er].height = 18
            is_sec = (rtype == "section")
            for ci_x, txt in enumerate([name, sym, unit], 1):
                c = ws.cell(er, ci_x, txt)
                c.border = brd; c.alignment = lft if ci_x==1 else ctr
                c.font = Font(size=10, bold=is_sec,
                              color="1A4A80" if is_sec else "222222")
                c.fill = fill("DCE8F5") if is_sec else fill("F5F8FC")
            for di in range(n_data):
                it  = self.table.item(ri, 3+di)
                raw = it.text() if it else ""
                try:    val = float(raw)
                except: val = raw
                c = ws.cell(er, 4+di, val)
                c.border = brd; c.alignment = ctr
                c.number_format = "0.0000" if isinstance(val,float) else "@"
                if is_sec:
                    c.fill=fill("DCE8F5"); c.font=Font(size=10,bold=True,color="1A4A80")
                elif rtype in ("coord",):
                    c.fill=fill("F5F8FC"); c.font=Font(size=10,color="555555")
                elif rtype == "calc" or (rtype=="rain_in" and self._city_mode):
                    if ri==_ROW_TOTAL and isinstance(val,float):
                        clr="0D47A1" if val<=15 else "880E4F"
                        c.fill=fill("E3F2FD" if val<=15 else "FCE4EC")
                        c.font=Font(size=10,bold=True,color=clr)
                    elif ri in(_ROW_RAIN,_ROW_GAS,_ROW_CLOUD) and isinstance(val,float):
                        clr="1A6B35" if val<=10 else "B05000"
                        c.fill=fill("E8F5E9" if val<=10 else "FFF3E0")
                        c.font=Font(size=10,bold=True,color=clr)
                    elif ri==_ROW_XPD_R and isinstance(val,float):
                        if val>=25:   c.fill=fill("E8F5E9"); c.font=Font(size=10,bold=True,color="1A6B35")
                        elif val>=10: c.fill=fill("FFF3E0"); c.font=Font(size=10,bold=True,color="B05000")
                        else:         c.fill=fill("FCE4EC"); c.font=Font(size=10,bold=True,color="880E4F")
                    else:
                        c.fill=fill("FFFBF0"); c.font=Font(size=10)
                else:
                    c.fill=fill("FFFFFF"); c.font=Font(size=10)

        ws.freeze_panes = "A3"
        wb.save(path)

    @staticmethod
    def _bstyle(color):
        from PyQt6.QtGui import QColor as QC
        dark = QC(color).darker(130).name()
        return (f"QPushButton{{background:{color};color:#FFF;border:none;"
                f"border-radius:4px;padding:0 12px;font-size:9pt;}}"
                f"QPushButton:hover{{background:{dark};}}")


# ══════════════════════════════════════════════════════════
#  主对话框
# ══════════════════════════════════════════════════════════

class ChannelModelDialog(ModuleDialog):
    TITLE        = "无线信道建模"
    ACCENT_COLOR = "#1D9E75"
    MIN_WIDTH    = 1100
    MIN_HEIGHT   = 700

    def build_content(self, layout: QVBoxLayout):
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(0)
        layout.addWidget(ChannelTableWidget(), stretch=1)